import os, sys
import openpyxl
import psycopg2

sys.stdout.reconfigure(encoding='utf-8')

env_vars = {}
if os.path.exists(".env"):
    with open(".env", "r", encoding="utf-8") as f:
        for line in f:
            if "=" in line and not line.startswith("#"):
                k, v = line.strip().split("=", 1)
                env_vars[k.strip()] = v.strip().strip('"').strip("'")

db_url = env_vars.get("DATABASE_URL") or env_vars.get("DIRECT_URL")
if not db_url:
    print("DATABASE_URL not found")
    exit(1)

# Clean query params for psycopg2
clean_url = db_url.split("?")[0]

print("Connecting to DB...")
conn = psycopg2.connect(clean_url)
cursor = conn.cursor()

tanmiya_path = r"C:\Users\my computer\Downloads\محدث كشف الأسر المستفيدة من مشروع كفالة الأسر المتعففة - 2025 - ناصر الدبوس تنمية (1)_٠٧٥٠٥٥.xlsx"
wb = openpyxl.load_workbook(tanmiya_path, data_only=True)
sheet = wb.active

updated_count = 0

for r in range(2, sheet.max_row + 1):
    name = str(sheet.cell(r, 2).value or "").strip()
    phone = str(sheet.cell(r, 5).value or "").strip()
    status_text = str(sheet.cell(r, 6).value or "").strip()

    if not name and not phone:
        continue

    is_displaced = "نازح" in status_text or "نازحه" in status_text
    is_widow = "أرمل" in status_text or "ارمل" in status_text
    is_severe = "فقيرة" in status_text or "صعبة" in status_text or "تعفف" in status_text

    clean_name_prefix = name[:8] if name else ""
    clean_phone = phone[-7:] if phone and len(phone) >= 7 else phone

    cursor.execute("""
        UPDATE families
        SET 
            "socialStatus" = CASE WHEN %s <> '' THEN %s ELSE "socialStatus" END,
            "isDisplaced" = CASE WHEN %s THEN true ELSE "isDisplaced" END,
            "hasWidow" = CASE WHEN %s THEN true ELSE "hasWidow" END,
            "povertyLevel" = CASE WHEN %s THEN 'SEVERE' ELSE "povertyLevel" END
        WHERE "headFullName" LIKE %s OR ("headPhoneNumber" IS NOT NULL AND "headPhoneNumber" LIKE %s)
    """, (status_text, status_text, is_displaced, is_widow, is_severe, f"%{clean_name_prefix}%", f"%{clean_phone}%"))

    if cursor.rowcount > 0:
        updated_count += cursor.rowcount

conn.commit()
print(f"✅ Updated {updated_count} family records with social status!")

cursor.execute('SELECT COUNT(*) FROM families WHERE "isDisplaced" = true;')
displaced_cnt = cursor.fetchone()[0]

cursor.execute('SELECT COUNT(*) FROM families WHERE "hasWidow" = true;')
widows_cnt = cursor.fetchone()[0]

cursor.execute('SELECT COUNT(*) FROM families WHERE "povertyLevel" = \'SEVERE\';')
severe_cnt = cursor.fetchone()[0]

print(f"📊 New System Totals:")
print(f"  • Displaced Families: {displaced_cnt}")
print(f"  • Widow Households: {widows_cnt}")
print(f"  • Severe Poverty Households: {severe_cnt}")

cursor.close()
conn.close()
